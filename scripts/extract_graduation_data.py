"""从 Excel 毕业率计算器中提取数据，生成 JSON 数据文件。

用法：
    python scripts/extract_graduation_data.py

读取 data/temp/excel/*.xlsx，输出到 data/graduation/{流派名}.json。
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

# ─── 路径 ──────────────────────────────────────────────────────
_EXCEL_DIR = Path(__file__).resolve().parent.parent / "data" / "temp" / "excel"
_OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data" / "graduation"

# ─── 文件名 → 流派配置名 映射 ──────────────────────────────────
_FILE_TO_SCHOOL: dict[str, str] = {
    "鸣金虹": "鸣金·虹",
    "鸣金影": "鸣金·影",
    "裂石威": "裂石·威",
    "裂石钧": "裂石·钧",
    "牵丝玉": "牵丝·玉",
    "牵丝霖": "牵丝·霖",
    "牵丝翊": "牵丝·翊",
    "破竹尘": "破竹·尘",
    "破竹风": "破竹·风",
    "破竹鸢": "破竹·鸢",
    "破竹樽": "破竹·樽",
}


def _find_excel(school_prefix: str) -> Path | None:
    """根据流派前缀查找对应的 Excel 文件"""
    for f in _EXCEL_DIR.iterdir():
        if f.suffix == ".xlsx" and school_prefix in f.stem:
            return f
    return None


def _to_num(v: Any) -> float | int | str:
    """将单元格值转为 JSON 友好的类型"""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "" or s == "N/a":
        return 0
    # 尝试转数字
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


def _extract_skills(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, dict]:
    """从武学奇术 sheet 提取技能定义"""
    # R1 为表头
    headers: dict[int, str] = {}
    for col in range(2, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None:
            headers[col] = str(v).strip()

    skills: dict[str, dict] = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()
        if name == "" or name == "N/a":
            continue
        skill: dict[str, Any] = {}
        for col, header in headers.items():
            val = _to_num(ws.cell(row=row, column=col).value)
            skill[header] = val
        skills[name] = skill
    return skills


def _extract_buffs(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, dict]:
    """从增益 sheet 提取增益定义"""
    headers: dict[int, str] = {}
    for col in range(2, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None:
            headers[col] = str(v).strip()

    buffs: dict[str, dict] = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()
        if name == "" or name == "N/a":
            continue
        buff: dict[str, Any] = {}
        for col, header in headers.items():
            val = _to_num(ws.cell(row=row, column=col).value)
            buff[header] = val
        buffs[name] = buff
    return buffs


def _extract_rotation(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    skills: dict[str, dict],
) -> list[dict]:
    """从期望 sheet 提取技能轴（rotation）

    动态检测 R1 中 "技能" 表头所在列，然后：
    - 技能列：取技能名
    - 技能列左侧：如果有 "次数" 列则取 hits，否则默认 1
    - 技能列右侧：取所有增益列，过滤 "N/a" 和空值
    - type：从 skills dict 中查找 "类型"
    """
    # 1. 在 R1 中找 "技能" 列
    skill_col = None
    hits_col = None
    buff_cols: list[int] = []

    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        label = str(v).strip()
        if label == "技能":
            skill_col = col
        elif label == "次数":
            hits_col = col
        elif label.startswith("增益"):
            buff_cols.append(col)

    if skill_col is None:
        print("  WARNING: 未找到 '技能' 列，跳过旋转提取")
        return []

    # 如果没找到 "次数" 列，用技能列左侧第一列
    if hits_col is None:
        hits_col = skill_col - 1

    # 如果没有找到增益列，取技能列右侧所有列
    if not buff_cols:
        buff_cols = list(range(skill_col + 1, ws.max_column + 1))

    rotation: list[dict] = []
    for row in range(2, ws.max_row + 1):
        # 技能列的值（核心判断依据）
        skill_val = ws.cell(row=row, column=skill_col).value
        if skill_val is None:
            continue
        skill_name = str(skill_val).strip()
        if skill_name == "":
            continue

        # 次数
        hits_val = ws.cell(row=row, column=hits_col).value
        try:
            hits = int(float(str(hits_val))) if hits_val is not None else 1
        except (ValueError, TypeError):
            hits = 1

        # 增益
        buffs: list[str] = []
        for bc in buff_cols:
            bv = ws.cell(row=row, column=bc).value
            if bv is None:
                continue
            bs = str(bv).strip()
            if bs == "" or bs == "N/a":
                continue
            buffs.append(bs)

        # 类型：从 skills dict 查找
        skill_type = ""
        if skill_name in skills:
            skill_type = str(skills[skill_name].get("类型", ""))
        elif skill_name == "N/a":
            # N/a 行尝试从增益推断（如 "心法"）
            # 在原始 JSON 中，N/a 行有 "心法" 类型
            # 检查增益列中是否有 "无惧" 等心法相关增益
            # 暂时标记为空，后续可调整
            skill_type = ""

        rotation.append({
            "row": row,
            "skill": skill_name,
            "hits": hits,
            "type": skill_type,
            "buffs": buffs,
        })

    return rotation


def _extract_metadata(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, float]:
    """从期望 sheet 提取元数据（战斗时间、baselineDps）"""
    metadata: dict[str, float] = {}

    for row in range(1, 30):
        for col in range(1, 10):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v).strip()
            if s == "战斗时间":
                # 战斗时间在同一行的 C9 或右侧某列
                for c in range(col + 1, 12):
                    tv = ws.cell(row=row, column=c).value
                    if tv is not None:
                        try:
                            metadata["combat_time"] = float(tv)
                            break
                        except (ValueError, TypeError):
                            pass
            elif s in ("DPS", "ADPS"):
                for c in range(col + 1, 12):
                    tv = ws.cell(row=row, column=c).value
                    if tv is not None:
                        try:
                            metadata["baseline_dps"] = float(tv)
                            break
                        except (ValueError, TypeError):
                            pass
    return metadata


def _infer_type_for_na(
    rotation: list[dict],
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> None:
    """为 N/a 行推断 type

    在原始鸣金虹 JSON 中，N/a 行有 "心法" 类型。
    检查期望 sheet 中 N/a 行附近是否有 "心法" 标记。
    """
    # 扫描期望 sheet，找 "心法" 标记
    xinfa_rows: set[int] = set()
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None and str(v).strip() == "心法":
                xinfa_rows.add(row)
                break

    if not xinfa_rows:
        return

    # 为 N/a 行设置类型
    for item in rotation:
        if item["skill"] == "N/a" and item["row"] in xinfa_rows:
            item["type"] = "心法"


def extract_one(xlsx_path: Path, school_name: str) -> dict[str, Any]:
    """从单个 Excel 文件提取完整数据"""
    print(f"  读取: {xlsx_path.name}")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    # 武学奇术
    if "武学奇术" not in wb.sheetnames:
        print(f"  WARNING: 未找到 '武学奇术' sheet")
        return {}
    skills = _extract_skills(wb["武学奇术"])
    print(f"    技能: {len(skills)} 个")

    # 增益
    if "增益" not in wb.sheetnames:
        print(f"  WARNING: 未找到 '增益' sheet")
        return {}
    buffs = _extract_buffs(wb["增益"])
    print(f"    增益: {len(buffs)} 个")

    # 旋转
    if "期望" not in wb.sheetnames:
        print(f"  WARNING: 未找到 '期望' sheet")
        return {}
    rotation = _extract_rotation(wb["期望"], skills)
    print(f"    技能轴: {len(rotation)} 条")

    # 为 N/a 行推断类型
    _infer_type_for_na(rotation, wb["期望"])

    # 元数据
    metadata = _extract_metadata(wb["期望"])
    print(f"    元数据: {metadata}")

    return {
        "rotation": rotation,
        "skills": skills,
        "buffs": buffs,
        "metadata": metadata,
    }


def main() -> None:
    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Excel 目录: {_EXCEL_DIR}")
    print(f"输出目录: {_OUTPUT_DIR}")
    print()

    success = 0
    failed = 0

    for prefix, school_name in sorted(_FILE_TO_SCHOOL.items(), key=lambda x: x[1]):
        xlsx = _find_excel(prefix)
        if xlsx is None:
            print(f"[SKIP] {school_name}: 未找到 Excel 文件 (前缀: {prefix})")
            failed += 1
            continue

        data = extract_one(xlsx, school_name)
        if not data:
            print(f"[FAIL] {school_name}: 提取失败")
            failed += 1
            continue

        out_path = _OUTPUT_DIR / f"{school_name}.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"[OK] {school_name} -> {out_path.name}")
        success += 1
        print()

    print(f"完成: {success} 成功, {failed} 失败/跳过")


if __name__ == "__main__":
    main()
