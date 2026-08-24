"""技能轴解析

毕业率方案 JSON 在编译时丢掉了技能名，看「伤害来自哪个技能」只能回到
原始 Excel。本模块只读解析「期望」工作表。

用例自己构造工作簿，不依赖任何本机 Excel——真实 11 份表的对账在
test_rotation_real_workbooks.py 里做，那批用例在文件缺失时自动跳过。
"""

import openpyxl
import pytest

from lvjiang.apps.yysls.core.graduation.rotation import (
    RotationParseError,
    parse_rotation,
)

# 真实表的列位置并不一致（技能列出现在第 17 和第 18 列），
# 因此这里刻意把必需列放在“奇怪”的位置，确保解析器按表头名定位。
_HEADERS = {
    3: "战斗时间",
    5: "期望",
    9: "次数",
    14: "技能",
    20: "类型",
}


def _make_book(path, rows, *, combat_time=100.0, sheet="期望", headers=None):
    """rows: [(技能, 次数, 期望, 类型), ...]"""
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = sheet
    for col, name in (headers or _HEADERS).items():
        ws.cell(1, col, name)
    if combat_time is not None:
        ws.cell(2, 3, "战斗时间")
        ws.cell(2, 4, combat_time)
    for i, (skill, count, damage, kind) in enumerate(rows, start=2):
        ws.cell(i, 14, skill)
        ws.cell(i, 9, count)
        ws.cell(i, 5, damage)
        ws.cell(i, 20, kind)
    book.save(path)
    return path


@pytest.fixture
def simple(tmp_path):
    return _make_book(tmp_path / "rot.xlsx", [
        ("三剑气", 2, 500.0, "剑"),
        ("飞剑", 1, 100.0, "剑"),
        ("三剑气", 1, 250.0, "剑"),
        ("心法切换", 1, 0.0, "心法"),
    ])


class TestParse:
    def test_reads_rows_in_axis_order(self, simple):
        rot = parse_rotation(simple)
        assert [h.skill for h in rot.hits] == ["三剑气", "飞剑", "三剑气", "心法切换"]
        assert [h.index for h in rot.hits] == [1, 2, 3, 4]

    def test_totals(self, simple):
        rot = parse_rotation(simple)
        assert rot.total_damage == pytest.approx(850.0)
        assert rot.total_hits == 5          # 2+1+1+1
        assert rot.combat_time == pytest.approx(100.0)
        assert rot.dps == pytest.approx(8.5)

    def test_damage_column_already_includes_count(self, simple):
        """期望列内部已含次数，聚合时不能再乘一遍。"""
        rot = parse_rotation(simple)
        assert sum(h.damage for h in rot.hits) == rot.total_damage

    def test_locates_columns_by_header_name(self, tmp_path):
        """列位置换一套照样解析——真实 11 份表的列号就不一致。"""
        moved = {1: "技能", 2: "次数", 3: "期望", 4: "类型"}
        path = _make_book(tmp_path / "m.xlsx", [], headers=moved, combat_time=None)
        book = openpyxl.load_workbook(path)
        ws = book["期望"]
        ws.cell(2, 1, "甲"), ws.cell(2, 2, 3), ws.cell(2, 3, 90.0), ws.cell(2, 4, "剑")
        book.save(path)
        rot = parse_rotation(path)
        assert rot.hits[0].skill == "甲"
        assert rot.hits[0].count == 3
        assert rot.hits[0].damage == pytest.approx(90.0)

    def test_blank_skill_rows_skipped(self, tmp_path):
        path = _make_book(tmp_path / "b.xlsx", [("甲", 1, 10.0, "剑")])
        book = openpyxl.load_workbook(path)
        ws = book["期望"]
        ws.cell(9, 9, 5)                    # 只有次数没有技能名 → 不是轴行
        ws.cell(10, 14, "   ")              # 空白技能名
        book.save(path)
        assert len(parse_rotation(path).hits) == 1

    def test_missing_count_defaults_to_one(self, tmp_path):
        path = _make_book(tmp_path / "c.xlsx", [("甲", None, 10.0, "剑")])
        assert parse_rotation(path).hits[0].count == 1

    def test_optional_columns_absent(self, tmp_path):
        """5 份真实表没有部分可选列，缺列不能报错。"""
        path = _make_book(tmp_path / "d.xlsx", [],
                          headers={1: "技能", 2: "次数", 3: "期望"},
                          combat_time=None)
        book = openpyxl.load_workbook(path)
        ws = book["期望"]
        ws.cell(2, 1, "甲"), ws.cell(2, 2, 1), ws.cell(2, 3, 10.0)
        book.save(path)
        rot = parse_rotation(path)
        assert rot.hits[0].kind == ""
        assert rot.combat_time == 0.0
        assert rot.dps == 0.0               # 没有时长时不能除零


class TestAggregate:
    def test_by_skill_merges_and_sorts(self, simple):
        by = parse_rotation(simple).by_skill()
        assert [s.skill for s in by] == ["三剑气", "飞剑", "心法切换"]
        top = by[0]
        assert top.rows == 2 and top.hits == 3
        assert top.damage == pytest.approx(750.0)
        assert top.share == pytest.approx(750 / 850 * 100)
        assert top.average == pytest.approx(250.0)

    def test_zero_damage_skill_sorted_last(self, simple):
        by = parse_rotation(simple).by_skill()
        assert by[-1].skill == "心法切换"
        assert by[-1].average == 0.0        # 不能除零

    def test_shares_sum_to_100(self, simple):
        by = parse_rotation(simple).by_skill()
        assert sum(s.share for s in by) == pytest.approx(100.0)


class TestErrors:
    def test_missing_sheet(self, tmp_path):
        path = tmp_path / "x.xlsx"
        book = openpyxl.Workbook()
        book.active.title = "别的表"
        book.save(path)
        with pytest.raises(RotationParseError, match="期望"):
            parse_rotation(path)

    def test_missing_required_column(self, tmp_path):
        path = _make_book(tmp_path / "y.xlsx", [],
                          headers={1: "技能", 2: "次数"}, combat_time=None)
        with pytest.raises(RotationParseError, match="期望"):
            parse_rotation(path)

    def test_no_skill_rows(self, tmp_path):
        path = _make_book(tmp_path / "z.xlsx", [])
        with pytest.raises(RotationParseError, match="技能行"):
            parse_rotation(path)

    def test_not_an_excel(self, tmp_path):
        path = tmp_path / "n.xlsx"
        path.write_text("不是 xlsx", encoding="utf-8")
        with pytest.raises(RotationParseError):
            parse_rotation(path)
