#!/usr/bin/env python
"""从毕业率计算器 Excel 抽出技能系数表与增益表。

毕业率方案 JSON 里的 ``program`` 能精确复现整张表，却是两千多个节点的
表达式图——读不出「第一道剑气的外功倍率是 1.3066」，也改不了。本脚本
把 `武学奇术` 与 `增益` 两张工作表抽成可读的 YAML，供「属性配置 →
伤害建模」展示与编辑。

只抽常量。Excel 里写成公式的格子（`=SUM(Z2:Z4)`、`=198.72%*2`）用
缓存值，缓存值缺失就报错退出——猜一个数字进配置比缺一条更糟。

环境参数（怪物防御/抗性、食物、固伤加成）不抽：方案 JSON 的
``environment`` 已经存了一份，存两份必然漏同步。

用法::

    .venv/bin/python scripts/extract_damage_model.py 鸣金·虹
    .venv/bin/python scripts/extract_damage_model.py 鸣金·虹 --dry-run
"""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]
EXCEL_DIR = ROOT / "data" / "excel"
SCHEME_DIR = ROOT / "config" / "system" / "yysls" / "graduation"
TARGET_DIR = ROOT / "config" / "system" / "yysls" / "damage_model"

#: `武学奇术` / `增益` 两张表的表头名 → 配置字段。两张表用同一套词汇，
#: 因为它们在求值里走的就是同一条加法。
MODIFIERS = {
    "通用增伤": "generic", "特殊增伤": "special",
    "最小外功": "min_outer", "最大外功": "max_outer",
    "外功加成": "outer_bonus", "外功穿透": "outer_pen",
    "外攻伤害加成": "outer_dmg", "属性攻击加成": "attr_bonus",
    "鸣金穿透": "mingjin_pen", "鸣金加成": "mingjin_dmg",
    "裂石穿透": "lieshi_pen", "裂石加成": "lieshi_dmg",
    "牵丝穿透": "qiansi_pen", "牵丝加成": "qiansi_dmg",
    "破竹穿透": "pozhu_pen", "破竹加成": "pozhu_dmg",
    "会心率": "crit_rate", "会心伤害": "crit_dmg",
    "会意率": "intent_rate", "会意伤害": "intent_dmg",
    "直接会心率": "direct_crit", "直接会意率": "direct_intent",
}
RATIOS = {
    "外功倍率": "outer_ratio", "外攻固伤": "outer_fixed",
    "属性倍率": "attr_ratio", "属性固伤": "attr_fixed",
}
FORCE = {
    "强制精准": "force_precision", "强制会心": "force_crit",
    "强制会意": "force_intent",
}


class ExtractError(RuntimeError):
    """表结构对不上。11 份表的列位置并不统一，只有表头名是。"""


def _headers(sheet) -> dict[str, int]:
    """表头名 → 列下标（1 基）。必须按名定位：列位置各表不一致。"""
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }


def _value(sheet, cached, row: int, column: int | None):
    """取格子的值。写成公式的走缓存值，缓存缺失即报错。"""
    if column is None:
        return None
    raw = sheet.cell(row, column).value
    if isinstance(raw, str) and raw.startswith("="):
        raw = cached.cell(row, column).value
        if raw is None:
            address = sheet.cell(row, column).coordinate
            raise ExtractError(f"{sheet.title}!{address} 是公式且没有缓存值")
    return raw


def _number(value) -> float:
    return round(float(value), 10) if isinstance(value, (int, float)) else 0.0


def _modifiers(sheet, cached, row: int, headers: dict[str, int]) -> dict[str, float]:
    out = {}
    for label, field in MODIFIERS.items():
        value = _number(_value(sheet, cached, row, headers.get(label)))
        if value:
            out[field] = value
    return out


def _skills(book, cached_book) -> dict[str, dict]:
    sheet, cached = book["武学奇术"], cached_book["武学奇术"]
    headers = _headers(sheet)
    missing = set(RATIOS) - set(headers)
    if missing:
        raise ExtractError(f"武学奇术 缺少列: {'、'.join(sorted(missing))}")
    skills: dict[str, dict] = {}
    for row in range(2, sheet.max_row + 1):
        name = str(sheet.cell(row, 1).value or "").strip()
        if not name or name == "N/a":
            continue
        entry: dict = {}
        kind = _value(sheet, cached, row, headers.get("类型"))
        if kind:
            entry["kind"] = str(kind).strip()
        if str(_value(sheet, cached, row, headers.get("定音加成")) or "") == "蓄力技":
            entry["charge"] = True
        qi = _number(_value(sheet, cached, row, headers.get("真气比例")))
        if qi:
            entry["qi_ratio"] = qi
        for label, field in RATIOS.items():
            value = _number(_value(sheet, cached, row, headers[label]))
            if value:
                entry[field] = value
        force = {
            field: True
            for label, field in FORCE.items()
            if _number(_value(sheet, cached, row, headers.get(label)))
        }
        if force:
            entry["force"] = force
        modifiers = _modifiers(sheet, cached, row, headers)
        if modifiers:
            entry["modifiers"] = modifiers
        skills[name] = entry
    return skills


def _buffs(book, cached_book) -> dict[str, dict]:
    sheet, cached = book["增益"], cached_book["增益"]
    headers = _headers(sheet)
    buffs: dict[str, dict] = {}
    for row in range(2, sheet.max_row + 1):
        name = str(sheet.cell(row, 1).value or "").strip()
        if not name or name == "N/a":
            continue
        # 全 0 的增益也要留：表里有它，轴上就可能挂它，缺了会显得
        # 「这个增益没建模」，而实际是它确实不加静态属性。
        buffs[name] = _modifiers(sheet, cached, row, headers)
    return buffs


def _scheme(school: str) -> tuple[str, dict]:
    """找到该流派的方案 JSON。伤害模型与它同源，source 要对得上。"""
    matches = sorted(SCHEME_DIR.glob(f"{school}_*.json"))
    if not matches:
        raise ExtractError(f"找不到 {school} 的毕业率方案，请先导入 Excel")
    path = matches[0]
    return path.stem[len(school) + 1:], json.loads(path.read_text(encoding="utf-8"))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("school", help="流派名，如 鸣金·虹")
    parser.add_argument("--dry-run", action="store_true", help="只打印，不写文件")
    args = parser.parse_args()

    scheme, data = _scheme(args.school)
    source = dict(data.get("source") or {})
    excel = EXCEL_DIR / str(source.get("file") or "")
    if not excel.exists():
        raise ExtractError(f"方案记录的 Excel 不在 data/excel 下: {excel.name}")
    digest = hashlib.sha256(excel.read_bytes()).hexdigest()
    if source.get("sha256") and digest != source["sha256"]:
        raise ExtractError(
            f"{excel.name} 的 sha256 与方案记录不一致，方案可能是从另一份表导入的")

    book = openpyxl.load_workbook(excel, data_only=False)
    cached_book = openpyxl.load_workbook(excel, data_only=True)
    document = {
        "school": args.school,
        "scheme": scheme,
        "source": {"file": excel.name, "version": source.get("version", ""),
                   "sha256": digest},
        "skills": _skills(book, cached_book),
        "buffs": _buffs(book, cached_book),
    }
    body = yaml.dump(document, allow_unicode=True, sort_keys=False, width=1000)
    header = (
        f"# 伤害建模: {args.school}\n"
        f"# 由 scripts/extract_damage_model.py 从 {excel.name} 抽取，勿手工重排\n"
        "#\n"
        "# 只存那份编译程序里读不出来的东西：技能系数与增益。怪物、食物、\n"
        "# 固伤加成等环境参数在 graduation/ 的方案 JSON 里，不在这里重复。\n"
    )
    print(f"{args.school}: 技能 {len(document['skills'])} 条，"
          f"增益 {len(document['buffs'])} 条")
    if args.dry_run:
        print(body)
        return 0
    TARGET_DIR.mkdir(parents=True, exist_ok=True)
    path = TARGET_DIR / f"{args.school}.yaml"
    path.write_text(header + body, encoding="utf-8")
    print(f"已写入 {path.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except ExtractError as exc:
        print(f"抽取失败: {exc}", file=sys.stderr)
        sys.exit(1)
