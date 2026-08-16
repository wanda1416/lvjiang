"""Convert graduation workbooks into executable, named JSON schemes."""
from __future__ import annotations

import hashlib
import json
import os
import re
import tempfile
from pathlib import Path
from typing import Any

import openpyxl
import yaml
from openpyxl.worksheet.formula import ArrayFormula

from lvjiang.constants import PROJECT_ROOT

from .excel_formula import FormulaError, FormulaModel, parse_formula
from .graduation_program import ProgramCompiler, ProgramRuntime

GRADUATION_DIR = PROJECT_ROOT / "config" / "system" / "yysls" / "graduation"
GAME_CONFIG_PATH = PROJECT_ROOT / "config" / "system" / "yysls" / "game_config.yaml"

INPUTS = {
    "min_outer": "期望!B2", "max_outer": "期望!C2",
    "outer_pen": "期望!D2", "outer_bonus": "期望!E2",
    "min_mingjin": "期望!B3", "max_mingjin": "期望!C3",
    "mingjin_pen": "期望!D3", "mingjin_bonus": "期望!E3",
    "min_lieshi": "期望!B4", "max_lieshi": "期望!C4",
    "lieshi_pen": "期望!D4", "lieshi_bonus": "期望!E4",
    "min_qiansi": "期望!B5", "max_qiansi": "期望!C5",
    "qiansi_pen": "期望!D5", "qiansi_bonus": "期望!E5",
    "min_pozhu": "期望!B6", "max_pozhu": "期望!C6",
    "pozhu_pen": "期望!D6", "pozhu_bonus": "期望!E6",
    "min_wuxiang": "期望!B7", "max_wuxiang": "期望!C7",
    "precision": "期望!B8", "crit_rate": "期望!B9",
    "intent_rate": "期望!B10", "direct_crit": "期望!B11",
    "direct_intent": "期望!B12", "crit_dmg": "期望!B13",
    "intent_dmg": "期望!B14", "all_skill_bonus": "期望!B15",
    "boss_bonus": "期望!B16", "weapon_bonus_primary": "期望!B17",
    "weapon_bonus_secondary": "期望!B18", "single_qs_bonus": "期望!B19",
    "group_qs_bonus": "期望!B20", "special_bonus": "期望!B21",
}
RATIO_INPUTS = {
    "outer_bonus", "mingjin_bonus", "lieshi_bonus", "qiansi_bonus",
    "pozhu_bonus", "precision", "crit_rate", "intent_rate", "direct_crit",
    "direct_intent", "crit_dmg", "intent_dmg", "all_skill_bonus",
    "boss_bonus", "single_qs_bonus", "group_qs_bonus",
}
OUTPUTS = {
    "combat_time": "期望!I8", "total_damage": "期望!I10",
    "dps": "期望!I12", "rdps": "期望!I14",
    "graduation_rate": "期望!I16",
}


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _version(filename: str) -> str:
    match = re.search(r"(\d+\.\d+)\.xlsx$", filename)
    return match.group(1) if match else "unknown"


def _affix_input_names(workbook, school: str) -> dict[str, list[str]]:
    config = yaml.safe_load(GAME_CONFIG_PATH.read_text(encoding="utf-8"))
    labels = {
        "all_skill_bonus": workbook["期望"]["A15"].value,
        "boss_bonus": workbook["期望"]["A16"].value,
        "weapon_bonus_primary": workbook["期望"]["A17"].value,
        "weapon_bonus_secondary": workbook["期望"]["A18"].value,
        "single_qs_bonus": workbook["期望"]["A19"].value,
        "group_qs_bonus": workbook["期望"]["A20"].value,
        "special_bonus": workbook["期望"]["A21"].value,
    }
    canonical_names: set[str] = set()
    for category in (config.get("affix_caps") or {}).values():
        raw = category.get("_aliases", []) if isinstance(category, dict) else []
        groups = raw.values() if isinstance(raw, dict) else [raw]
        for names in groups:
            if isinstance(names, list):
                canonical_names.update(str(name) for name in names)
    alias_index: dict[str, list[str]] = {}
    for affix_name, aliases in (config.get("affix_aliases") or {}).items():
        if affix_name not in canonical_names:
            raise RuntimeError(f"affix_aliases uses unknown affix: {affix_name!r}")
        if not isinstance(aliases, list):
            raise RuntimeError(f"affix_aliases[{affix_name!r}] must be a list")
        for alias in aliases:
            alias_index.setdefault(str(alias), []).append(str(affix_name))
    skill_category = (config.get("affix_caps") or {}).get("指定技能增效") or {}
    skill_groups = skill_category.get("_aliases") if isinstance(
        skill_category, dict,
    ) else None
    if not isinstance(skill_groups, dict):
        raise RuntimeError(
            "游戏配置缺少 affix_caps.指定技能增效._aliases 流派分组，"
            "无法解析 Excel 的指定技能增效。请补充配置后重新导入。"
        )
    school_skill_names = skill_groups.get(school)
    if not isinstance(school_skill_names, list) or not school_skill_names:
        raise RuntimeError(
            f"游戏配置没有流派「{school}」的指定技能增效分组，无法解析 Excel。"
            "请补充流派配置后重新导入。"
        )
    school_config = (config.get("schools") or {}).get(school)
    if not isinstance(school_config, dict):
        raise RuntimeError(f"游戏配置缺少流派「{school}」定义，无法筛选 Excel 词条。")
    weapon_affixes = {
        str(entry.get("name")): str(entry.get("wuxue_affix"))
        for entry in (config.get("weapon_types") or [])
        if isinstance(entry, dict) and entry.get("name") and entry.get("wuxue_affix")
    }
    school_weapon_names = {
        weapon_affixes.get(str((school_config.get(position) or {}).get("weapon")))
        for position in ("main", "sub")
    } - {None}
    category_names = {
        name: set((config.get("affix_caps", {}).get(name) or {}).get("_aliases") or [])
        for name in ("全部武学增效", "对单位增效", "奇术类增伤")
    }
    allowed_by_input = {
        "all_skill_bonus": category_names["全部武学增效"],
        "boss_bonus": category_names["对单位增效"],
        "weapon_bonus_primary": school_weapon_names,
        "weapon_bonus_secondary": school_weapon_names,
        "single_qs_bonus": category_names["奇术类增伤"],
        "group_qs_bonus": category_names["奇术类增伤"],
    }
    result: dict[str, list[str]] = {}
    for input_name, label in labels.items():
        if label is None or not str(label).strip():
            if input_name == "special_bonus":
                # 部分流派没有指定技能增效；Excel 留空即明确表示不参与计算。
                result[input_name] = []
                continue
            raise RuntimeError(
                f"{school} Excel 输入 {input_name} 的字段名为空。"
                "请修正 Excel，或在词组配置中补充对应别名后重新导入。"
            )
        if input_name == "special_bonus":
            # 只在当前流派的指定技能增效分组中反查，按配置顺序取第一个。
            matches = [
                str(affix_name) for affix_name in school_skill_names
                if str(label) in (config.get("affix_aliases") or {}).get(
                    str(affix_name), [],
                )
            ]
            if not matches:
                raise RuntimeError(
                    f"流派「{school}」无法解析 Excel 指定技能增效简称 "
                    f"{label!r}。请在该流派的指定技能增效词条中配置此别名，"
                    "或修正 Excel 字段名后重新导入。"
                )
            result[input_name] = [matches[0]]
            continue
        matches = [
            name for name in alias_index.get(str(label), [])
            if name in allowed_by_input[input_name]
        ]
        if len(matches) != 1:
            raise RuntimeError(
                f"流派「{school}」的 Excel 字段简称 {label!r} 应唯一映射到"
                f"一个精准词条，实际匹配 {matches!r}。请修改 Excel 或词组别名"
                "配置后重新导入。"
            )
        result[input_name] = matches
    return result


def _cell_value(formulas, cached, address: str) -> Any:
    sheet, coordinate = address.split("!", 1)
    cached_value = cached[sheet][coordinate].value
    return _json_value(cached_value if cached_value is not None else formulas[sheet][coordinate].value)


def _extract_environment(formulas, cached) -> dict[str, Any]:
    sheet = formulas["期望"]
    cached_sheet = cached["期望"]
    team_buffs = []
    for row in range(16, 21):
        for name_col, enabled_col in ((3, 4), (5, 6)):
            name = sheet.cell(row, name_col).value
            if name not in (None, ""):
                team_buffs.append({
                    "name": str(name),
                    "enabled": str(sheet.cell(row, enabled_col).value or "") == "√",
                })
    monster = {}
    for row in range(2, 6):
        name = sheet.cell(row, 8).value
        if name not in (None, ""):
            value = cached_sheet.cell(row, 9).value
            if value is None:
                value = sheet.cell(row, 9).value
            monster[str(name)] = _json_value(value)
    return {
        "food_bonus": {
            "min_outer": _cell_value(formulas, cached, "期望!B23"),
            "max_outer": _cell_value(formulas, cached, "期望!B24"),
        },
        "fixed_damage_bonus": _cell_value(formulas, cached, "期望!B22"),
        "team_buffs": team_buffs,
        "monster": monster,
        "combat_time": _cell_value(formulas, cached, "期望!I8"),
    }


def _baseline_attrs(model: dict[str, Any], affix_names: dict[str, list[str]]) -> dict[str, Any]:
    runtime = FormulaModel(model)
    attrs: dict[str, Any] = {}
    extra_attrs: dict[str, float] = {}
    dynamic = {"weapon_bonus_primary", "weapon_bonus_secondary", "special_bonus"}
    for name, target in INPUTS.items():
        value = float(runtime.value(target) or 0)
        if name in dynamic:
            names = affix_names[name]
            if names:
                extra_attrs[names[0]] = value
        else:
            attrs[name] = value
    if extra_attrs:
        attrs["extra_attrs"] = extra_attrs
    return attrs


def _compile_v2(
    workbook_model: dict[str, Any], school: str, affix_names: dict[str, list[str]],
    environment: dict[str, Any],
) -> dict[str, Any]:
    dynamic = {"weapon_bonus_primary", "weapon_bonus_secondary", "special_bonus"}
    bindings: dict[str, dict[str, str]] = {}
    for name, target in INPUTS.items():
        if name in dynamic:
            names = affix_names[name]
            if not names:
                # A deliberately empty skill bonus remains a fixed zero.
                continue
            bindings[target] = {"kind": "affix", "name": names[0]}
        else:
            bindings[target] = {"kind": "field", "name": name}
    compiler = ProgramCompiler(workbook_model, bindings)
    program = compiler.compile({
        name: OUTPUTS[name]
        for name in ("combat_time", "total_damage", "dps", "graduation_rate")
    })
    baseline_attrs = _baseline_attrs(workbook_model, affix_names)
    input_values = []
    extra = baseline_attrs.get("extra_attrs", {})
    for spec in program["inputs"]:
        input_values.append(float(
            baseline_attrs.get(spec["name"], 0)
            if spec["kind"] == "field" else extra.get(spec["name"], 0)
        ))
    compiled_outputs = ProgramRuntime(program, input_values).outputs()
    workbook_runtime = FormulaModel(workbook_model)
    reference = {
        name: float(workbook_runtime.value(OUTPUTS[name]))
        for name in compiled_outputs
    }
    for name, actual in reference.items():
        sheet, coordinate = OUTPUTS[name].split("!", 1)
        cached = workbook_model["sheets"][sheet]["cells"][coordinate].get("cached")
        if cached is not None and abs(actual - float(cached)) > max(
            1e-6, abs(float(cached)) * 1e-10,
        ):
            raise FormulaError(
                f"Excel 公式 {name}={actual} 与工作簿缓存值 {cached} 不一致"
            )
    for name, actual in compiled_outputs.items():
        expected = reference[name]
        if abs(actual - expected) > max(1e-6, abs(expected) * 1e-10):
            raise FormulaError(
                f"编译后的 {name}={actual} 与 Excel 公式结果 {expected} 不一致"
            )
    return {
        "schema_version": 2,
        "school": school,
        "source": workbook_model["model"]["source"],
        "baseline_attrs": baseline_attrs,
        "environment": environment,
        "reference": reference,
        "program": program,
    }


def convert_workbook(path: Path, school: str) -> dict[str, Any]:
    formulas = openpyxl.load_workbook(path, data_only=False, read_only=False)
    cached = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        affix_names = _affix_input_names(formulas, school)
        sheets: dict[str, Any] = {}
        formula_count = 0
        for worksheet in formulas.worksheets:
            cached_sheet = cached[worksheet.title]
            cells: dict[str, Any] = {}
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    formula = value.text if isinstance(value, ArrayFormula) else value
                    if isinstance(formula, str) and formula.startswith("="):
                        parse_formula(formula)
                        entry: dict[str, Any] = {"formula": formula}
                        cached_value = cached_sheet[cell.coordinate].value
                        if cached_value is not None:
                            entry["cached"] = _json_value(cached_value)
                        formula_count += 1
                    else:
                        entry = {"value": _json_value(value)}
                    cells[cell.coordinate] = entry
            sheets[worksheet.title] = {
                "dimensions": {
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                },
                "cells": cells,
            }
        source_bytes = path.read_bytes()
        workbook_model = {
            "schema_version": "1.0",
            "formula_language": "excel_subset_v1",
            "model": {
                "id": school.replace("·", ""), "school": school,
                "source": {
                    "file": path.name, "version": _version(path.name),
                    "sha256": hashlib.sha256(source_bytes).hexdigest(),
                },
                "formula_count": formula_count,
            },
            "inputs": {
                name: {
                    "type": "number",
                    "unit": "ratio" if name in RATIO_INPUTS or name in {
                        "weapon_bonus_primary", "weapon_bonus_secondary",
                        "special_bonus",
                    } else "point",
                    "target": target,
                    **({"label_ref": f"期望!A{target.rsplit('B', 1)[-1]}",
                        "affix_names": affix_names[name]}
                       if name in affix_names else {}),
                }
                for name, target in INPUTS.items()
            },
            "outputs": {name: {"ref": ref} for name, ref in OUTPUTS.items()},
            "sheets": sheets,
        }
        return _compile_v2(
            workbook_model, school, affix_names,
            _extract_environment(formulas, cached),
        )
    finally:
        formulas.close()
        cached.close()


def validate_model(model: dict[str, Any]) -> dict[str, float]:
    if model.get("schema_version") != 2:
        raise FormulaError("graduation model must use schema version 2")
    baseline = model["baseline_attrs"]
    extra = baseline.get("extra_attrs", {})
    values = [
        float(baseline.get(spec["name"], 0) if spec["kind"] == "field"
              else extra.get(spec["name"], 0))
        for spec in model["program"]["inputs"]
    ]
    results = ProgramRuntime(model["program"], values).outputs()
    for name, expected in model["reference"].items():
        if abs(results[name] - float(expected)) > max(1e-6, abs(float(expected)) * 1e-10):
            raise FormulaError(f"compiled {name} does not match its reference value")
    return results


def scheme_path(school: str, scheme: str) -> Path:
    cleaned = scheme.strip()
    if not cleaned or any(char in cleaned for char in '<>:"/\\|?*'):
        raise ValueError("方案名称为空或包含文件名非法字符")
    return GRADUATION_DIR / f"{school}_{cleaned}.json"


def write_model(path: Path, model: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(model, ensure_ascii=False, indent=2) + "\n"
    fd, temp_name = tempfile.mkstemp(dir=path.parent, suffix=".json.tmp")
    temp_path = Path(temp_name)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as stream:
            stream.write(payload)
        os.replace(temp_path, path)
    except BaseException:
        temp_path.unlink(missing_ok=True)
        raise


def import_graduation_scheme(
    excel_path: str | Path, school: str, scheme: str,
) -> tuple[Path, dict[str, float]]:
    source = Path(excel_path)
    if source.suffix.lower() != ".xlsx" or not source.is_file():
        raise ValueError("请选择有效的 .xlsx 文件")
    model = convert_workbook(source, school)
    outputs = validate_model(model)
    destination = scheme_path(school, scheme)
    write_model(destination, model)
    return destination, outputs
