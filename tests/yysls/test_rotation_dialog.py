"""技能轴查看器对话框"""

import openpyxl
import pytest
from PyQt6.QtWidgets import QProgressBar

from lvjiang.apps.yysls.ui.loadout.rotation_dialog import RotationDialog


@pytest.fixture
def workbook(tmp_path):
    path = tmp_path / "rot.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "期望"
    for col, name in ((1, "技能"), (2, "次数"), (3, "期望"), (4, "类型")):
        ws.cell(1, col, name)
    ws.cell(2, 6, "战斗时间"), ws.cell(2, 7, 100.0)
    for row, (skill, cnt, dmg, kind) in enumerate([
        ("三剑气", 2, 600.0, "剑"),
        ("飞剑", 1, 300.0, "剑"),
        ("三剑气", 1, 100.0, "剑"),
        ("心法", 1, 0.0, "心法"),
    ], start=2):
        ws.cell(row, 1, skill), ws.cell(row, 2, cnt)
        ws.cell(row, 3, dmg), ws.cell(row, 4, kind)
    book.save(path)
    return path


class TestEmptyState:
    def test_opens_without_file(self, qtbot):
        dlg = RotationDialog()
        qtbot.addWidget(dlg)
        assert dlg.rotation is None
        assert dlg._table_axis.rowCount() == 0
        assert dlg._table_source.rowCount() == 0


class TestRender:
    @pytest.fixture
    def dlg(self, qtbot, workbook):
        d = RotationDialog(initial=workbook)
        qtbot.addWidget(d)
        return d

    def test_axis_rows_in_order(self, dlg):
        assert dlg._table_axis.rowCount() == 4
        assert [dlg._table_axis.item(r, 1).text() for r in range(4)] == [
            "三剑气", "飞剑", "三剑气", "心法"]

    def test_source_aggregated_and_sorted(self, dlg):
        assert [dlg._table_source.item(r, 0).text() for r in range(3)] == [
            "三剑气", "飞剑", "心法"]
        assert dlg._table_source.item(0, 4).text() == "700"     # 600+100
        assert dlg._table_source.item(0, 3).text() == "3"       # 命中 2+1

    def test_share_rendered_as_progress_bar(self, dlg):
        bar = dlg._table_source.cellWidget(0, 5)
        assert isinstance(bar, QProgressBar)
        assert bar.format() == "70.00%"

    def test_summary_line(self, dlg):
        text = dlg._lbl_summary.text()
        assert "4 行" in text and "5 次命中" in text and "3 个技能" in text
        assert "1,000" in text          # 总伤
        assert "10" in text             # DPS = 1000/100

    def test_zero_damage_row_has_hint(self, dlg):
        assert dlg._table_source.item(2, 0).toolTip()


class TestErrorHandling:
    def test_bad_file_warns_and_keeps_state(self, qtbot, tmp_path, monkeypatch):
        import lvjiang.apps.yysls.ui.loadout.rotation_dialog as mod
        warned = []
        monkeypatch.setattr(
            mod.QMessageBox, "warning",
            lambda *a, **k: warned.append(a[2] if len(a) > 2 else ""))
        bad = tmp_path / "bad.xlsx"
        bad.write_text("不是 xlsx", encoding="utf-8")
        dlg = RotationDialog(initial=bad)
        qtbot.addWidget(dlg)
        assert warned, "解析失败应提示用户"
        assert dlg.rotation is None
        assert dlg._table_axis.rowCount() == 0

    def test_reload_replaces_previous(self, qtbot, workbook, tmp_path):
        """再次导入要整表替换，不能把两条轴叠在一起。"""
        dlg = RotationDialog(initial=workbook)
        qtbot.addWidget(dlg)
        assert dlg._table_axis.rowCount() == 4
        dlg._load(workbook)
        assert dlg._table_axis.rowCount() == 4


class TestEntryPoint:
    """入口只在「最优组合」对话框首行最右侧——实验性功能，不进主工具栏

    技能轴需要用户自备毕业率计算器 Excel，且与备战流程无关，因此刻意不放到
    备战方案主工具栏上。曾经放过两处都不合适：EquipStatusTab 的 action_row
    在 LoadoutPanel 里被 set_embedded_mode(True) 整个隐藏（用户看不到），
    主工具栏又过于显眼。
    """

    @pytest.fixture
    def dialog(self, qtbot, tmp_path, monkeypatch):
        import lvjiang.constants
        from lvjiang.apps.yysls.core.combat.combat_attrs import CombatAttributes
        from lvjiang.apps.yysls.ui.loadout.optimal_combo import (
            OptimalComboDialog,
        )

        monkeypatch.setattr(lvjiang.constants, "USERS_DIR", tmp_path)
        monkeypatch.setattr(
            lvjiang.constants, "SESSION_PATH", tmp_path / "session.json")
        class _Host:
            def active_user_name(self):
                return ""      # 空用户名 → 跳过装备加载，够本用例用

        dlg = OptimalComboDialog(
            host=_Host(), school="鸣金·虹", scheme="基础方案",
            base_attrs=CombatAttributes(),
        )
        qtbot.addWidget(dlg)
        return dlg

    def test_button_on_first_row_rightmost(self, dialog):
        from PyQt6.QtWidgets import QHBoxLayout

        btn = dialog._btn_rotation
        assert btn.text() == "技能轴"
        # 找到承载它的那个水平行，确认它是最后一个控件
        holder = None
        for child in dialog.findChildren(QHBoxLayout):
            for i in range(child.count()):
                if child.itemAt(i).widget() is btn:
                    holder = child
        assert holder is not None, "技能轴按钮应在一个水平行里"
        assert holder.itemAt(holder.count() - 1).widget() is btn, "应在该行最右侧"

    def test_marked_as_experimental_in_tooltip(self, dialog):
        assert "实验性" in dialog._btn_rotation.toolTip()

    def test_click_opens_dialog(self, dialog, monkeypatch):
        opened = []
        monkeypatch.setattr(
            "lvjiang.apps.yysls.ui.loadout.rotation_dialog.RotationDialog.exec",
            lambda self: opened.append(self))
        dialog._btn_rotation.click()
        assert opened, "点击应打开技能轴对话框"

    def test_not_on_loadout_main_toolbar(self, qtbot, tmp_path, monkeypatch):
        """主工具栏上不应出现——实验性功能不直接开放。"""
        from PyQt6.QtCore import pyqtSignal
        from PyQt6.QtWidgets import QComboBox, QPushButton, QWidget

        import lvjiang.constants
        from lvjiang.apps.yysls.ui.loadout import LoadoutPanel

        monkeypatch.setattr(lvjiang.constants, "USERS_DIR", tmp_path)
        monkeypatch.setattr(
            lvjiang.constants, "SESSION_PATH", tmp_path / "session.json")

        class Host(QWidget):
            user_changed = pyqtSignal(str)
            equipment_changed = pyqtSignal()
            graduation_updated = pyqtSignal(object)
            open_play_style_form = pyqtSignal(dict)

            def __init__(self):
                super().__init__()
                self.user_combo = QComboBox(self)
                self.user_combo.addItems(["alice"])

            def active_user_name(self):
                return "alice"

            def navigate_user(self, _delta):
                pass

        host = Host()
        panel = LoadoutPanel(host)
        qtbot.addWidget(host)
        qtbot.addWidget(panel)
        assert not [b for b in panel.findChildren(QPushButton)
                    if b.text() == "技能轴"]
