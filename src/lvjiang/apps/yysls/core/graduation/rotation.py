"""技能轴解析 —— 从毕业率计算器 Excel 读取竞速轴

毕业率方案的 Excel（`*竞速轴属性毕业率进阶计算器*.xlsx`）里，「期望」工作表
就是竞速轴本体：一行一次技能释放，`期望` 列即该行的伤害贡献，全列求和恰好
等于方案的 ``reference.total_damage``。

编译成方案 JSON 时技能名会被丢弃（整个节点程序里只剩一个字符串常量），
所以要看「伤害来自哪个技能」只能回到 Excel。本模块只做只读解析，不改表、
不重算——伤害数值直接取 Excel 已算好的 `期望` 列。

**必须按表头名定位列**：11 份表的列位置并不一致（技能列在第 17 或 18 列，
类型列在 33/34/35/38/39 列都出现过），只有第 1 行的表头名是统一的。
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from .....i18n import tr

#: 轴所在的工作表名
SHEET_NAME = "期望"

#: 必需列的表头名 → 内部字段
_REQUIRED = {
    "技能": "skill",
    "次数": "count",
    "期望": "damage",
}
#: 可选列（部分流派的表没有）
_OPTIONAL = {
    "类型": "kind",
    "外功倍率": "outer_ratio",
    "属性倍率": "attr_ratio",
}


@dataclass(frozen=True)
class RotationHit:
    """轴上的一行：一次（或连续多次）技能释放"""
    index: int          # 轴序，从 1 开始
    skill: str
    count: int          # 该行代表几次命中
    kind: str           # 技能类型（剑/枪/群体奇术/心法…），缺列时为空
    damage: float       # 该行的伤害贡献；已含 count，不要再乘
    outer_ratio: float | None = None
    attr_ratio: float | None = None


@dataclass(frozen=True)
class SkillDamage:
    """按技能名聚合后的伤害来源"""
    skill: str
    kind: str
    rows: int
    hits: int
    damage: float
    share: float        # 占总伤百分比（0-100）

    @property
    def average(self) -> float:
        """单次命中均值；命中数为 0 时返回 0"""
        return self.damage / self.hits if self.hits else 0.0


@dataclass(frozen=True)
class Rotation:
    """一条完整的竞速轴"""
    source: str                      # 源 Excel 文件名
    hits: tuple[RotationHit, ...]
    total_damage: float
    combat_time: float               # 秒；表里没有时为 0

    @property
    def total_hits(self) -> int:
        return sum(h.count for h in self.hits)

    @property
    def dps(self) -> float:
        return self.total_damage / self.combat_time if self.combat_time else 0.0

    def by_skill(self) -> list[SkillDamage]:
        """按技能聚合，伤害降序。零伤害技能（心法切换等占位行）排在最后。"""
        buckets: dict[str, dict] = {}
        for hit in self.hits:
            b = buckets.setdefault(
                hit.skill, {"kind": hit.kind, "rows": 0, "hits": 0, "damage": 0.0})
            b["rows"] += 1
            b["hits"] += hit.count
            b["damage"] += hit.damage
            if not b["kind"]:
                b["kind"] = hit.kind
        total = self.total_damage or 1.0
        items = [
            SkillDamage(skill=name, kind=b["kind"], rows=b["rows"], hits=b["hits"],
                        damage=b["damage"], share=b["damage"] / total * 100)
            for name, b in buckets.items()
        ]
        items.sort(key=lambda s: (-s.damage, s.skill))
        return items


class RotationParseError(ValueError):
    """Excel 不是可识别的毕业率计算器，或缺少必需列"""


@dataclass
class _Header:
    """第 1 行表头名 → 列下标（0 基）"""
    columns: dict[str, int] = field(default_factory=dict)

    def index(self, name: str) -> int | None:
        return self.columns.get(name)


def _read_rows(path: Path) -> list[tuple]:
    """按行读出「期望」表

    用 ``iter_rows`` 而不是随机 ``cell(r, c)``：read_only 模式下随机访问会
    退化到秒级（实测最大的一份表 7.4s vs 按行 80ms）。
    """
    import openpyxl

    try:
        book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - 损坏文件统一转成可读错误
        raise RotationParseError(
            tr("无法打开 {name}：{err}").format(name=path.name, err=exc)) from exc
    try:
        if SHEET_NAME not in book.sheetnames:
            raise RotationParseError(
                tr("{name} 里没有「{sheet}」工作表，可能不是毕业率计算器")
                .format(name=path.name, sheet=SHEET_NAME))
        return list(book[SHEET_NAME].iter_rows(values_only=True))
    finally:
        book.close()


def _parse_header(row: tuple) -> _Header:
    header = _Header()
    for i, cell in enumerate(row):
        if isinstance(cell, str):
            header.columns.setdefault(cell.strip(), i)
    missing = [name for name in _REQUIRED if name not in header.columns]
    if missing:
        raise RotationParseError(
            tr("「{sheet}」表缺少必需列：{cols}")
            .format(sheet=SHEET_NAME, cols="、".join(missing)))
    return header


def _cell(row: tuple, index: int | None):
    if index is None or index >= len(row):
        return None
    return row[index]


def _number(value, default: float = 0.0) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return default
    return float(value)


def _combat_time(rows: list[tuple], header: _Header) -> float:
    """战斗时间在左侧属性面板里，按「战斗时间」标签横向找相邻数值"""
    for row in rows:
        for i, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip() == "战斗时间":
                for nxt in row[i + 1:i + 4]:
                    value = _number(nxt)
                    if value > 0:
                        return value
    return 0.0


def parse_rotation(path: str | Path) -> Rotation:
    """解析一份毕业率计算器 Excel，返回其竞速轴

    Raises:
        RotationParseError: 文件打不开、没有「期望」表、或缺少必需列
    """
    source = Path(path)
    rows = _read_rows(source)
    if not rows:
        raise RotationParseError(
            tr("「{sheet}」表是空的").format(sheet=SHEET_NAME))
    header = _parse_header(rows[0])

    col = {name: header.index(name) for name in {**_REQUIRED, **_OPTIONAL}}
    hits: list[RotationHit] = []
    for row in rows[1:]:
        raw_skill = _cell(row, col["技能"])
        if raw_skill is None or not str(raw_skill).strip():
            continue        # 轴之外的空行/汇总行
        count = int(_number(_cell(row, col["次数"]), 1.0)) or 1
        hits.append(RotationHit(
            index=len(hits) + 1,
            skill=str(raw_skill).strip(),
            count=count,
            kind=str(_cell(row, col["类型"]) or "").strip(),
            damage=_number(_cell(row, col["期望"])),
            outer_ratio=_cell(row, col["外功倍率"]) if isinstance(
                _cell(row, col["外功倍率"]), (int, float)) else None,
            attr_ratio=_cell(row, col["属性倍率"]) if isinstance(
                _cell(row, col["属性倍率"]), (int, float)) else None,
        ))
    if not hits:
        raise RotationParseError(
            tr("「{sheet}」表里没有找到任何技能行").format(sheet=SHEET_NAME))
    return Rotation(
        source=source.name,
        hits=tuple(hits),
        total_damage=sum(h.damage for h in hits),
        combat_time=_combat_time(rows, header),
    )
